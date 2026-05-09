from flask import Flask, request, jsonify, send_file
import json, os, random, string, hashlib, jwt, csv, io
from datetime import datetime, timedelta
from functools import wraps

# ── PostgreSQL or SQLite ─────────────────────────────────────────────────────
DATABASE_URL = os.environ.get('DATABASE_URL', '')
USE_PG = bool(DATABASE_URL)
if USE_PG:
    import psycopg2, psycopg2.extras
else:
    import sqlite3

DB_PATH = os.environ.get('DB_PATH', 'data/exam_system.db')

app = Flask(__name__, static_folder='public', static_url_path='/static')
app.secret_key = 'razi_exam_2024_secret'
JWT_SECRET   = 'razi_jwt_2024'

SCHOOL_CONFIG = {
    "name_ar":      "مدرسة الرازي حلقة ثانية بنين",
    "name_en":      "Al-Razi Boys School Cycle 2",
    "address":      "دبي، الامارات العربية المتحدة",
    "phone":        "",
    "email":        "",
    "copyright":    "جميع الحقوق محفوظة : مدرسة الرازي - مديرة المدرسة أ. بدرية سيف - تصميم : هاني ابوالدهب",
    "banner_color": "#1a3a5c",
    "ai_key":       ""
}

# ── DB Helpers ───────────────────────────────────────────────────────────────
def get_db():
    if USE_PG:
        conn = psycopg2.connect(DATABASE_URL)
        conn.autocommit = False
        return conn
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    return conn

def q(sql):
    return sql.replace('?','%s') if USE_PG else sql

def fetchall(conn, sql, p=()):
    c = conn.cursor(); c.execute(q(sql), p)
    if USE_PG:
        cols = [d[0] for d in c.description]
        return [dict(zip(cols,r)) for r in c.fetchall()]
    return [dict(r) for r in c.fetchall()]

def fetchone(conn, sql, p=()):
    c = conn.cursor(); c.execute(q(sql), p)
    if USE_PG:
        r = c.fetchone()
        if not r: return None
        return dict(zip([d[0] for d in c.description], r))
    r = c.fetchone()
    return dict(r) if r else None

def execute(conn, sql, p=()):
    c = conn.cursor(); c.execute(q(sql), p); return c

def insert(conn, sql, p=()):
    c = conn.cursor()
    if USE_PG:
        s = q(sql)
        if 'RETURNING' not in s.upper(): s += ' RETURNING id'
        c.execute(s, p)
        r = c.fetchone(); return r[0] if r else None
    c.execute(sql, p); return c.lastrowid

def hash_pw(p): return hashlib.sha256(p.encode()).hexdigest()
def gen_code(n=8): return ''.join(random.choices(string.ascii_uppercase+string.digits, k=n))
def grade_letter(pct):
    for threshold, letter in [(95,'A+'),(90,'A'),(85,'B+'),(80,'B'),(75,'C+'),(70,'C'),(65,'D+'),(60,'D')]:
        if pct >= threshold: return letter
    return 'F'

# ── DB Init ───────────────────────────────────────────────────────────────────
def init_db():
    os.makedirs('data', exist_ok=True)
    os.makedirs('public/images', exist_ok=True)
    conn = get_db()
    c = conn.cursor()
    AUTO = 'SERIAL PRIMARY KEY' if USE_PG else 'INTEGER PRIMARY KEY AUTOINCREMENT'
    IGNORE = 'ON CONFLICT DO NOTHING' if USE_PG else 'OR IGNORE'
    REPLACE = 'ON CONFLICT(key) DO NOTHING' if USE_PG else 'OR IGNORE'

    tables = [
        f'''CREATE TABLE IF NOT EXISTS school_settings (key TEXT PRIMARY KEY, value TEXT)''',
        f'''CREATE TABLE IF NOT EXISTS users (
            id {AUTO}, username TEXT UNIQUE NOT NULL, password_hash TEXT NOT NULL,
            role TEXT NOT NULL, full_name_ar TEXT, full_name_en TEXT, email TEXT,
            student_code TEXT UNIQUE, class_name TEXT, grade TEXT,
            created_at TEXT DEFAULT 'now')''',
        f'''CREATE TABLE IF NOT EXISTS subjects (
            id {AUTO}, name_ar TEXT NOT NULL, name_en TEXT, code TEXT UNIQUE NOT NULL,
            teacher_id INTEGER, grade TEXT, color TEXT DEFAULT '#3b82f6', icon TEXT DEFAULT '📚')''',
        f'''CREATE TABLE IF NOT EXISTS question_bank (
            id {AUTO}, subject_id INTEGER NOT NULL, question_ar TEXT NOT NULL,
            question_en TEXT, type TEXT NOT NULL, options_ar TEXT, options_en TEXT,
            correct_answer TEXT, points REAL DEFAULT 1, difficulty TEXT DEFAULT 'medium',
            skill_tag TEXT, created_by INTEGER, created_at TEXT DEFAULT 'now')''',
        f'''CREATE TABLE IF NOT EXISTS exams (
            id {AUTO}, title_ar TEXT NOT NULL, title_en TEXT, subject_id INTEGER NOT NULL,
            teacher_id INTEGER NOT NULL, instructions_ar TEXT, instructions_en TEXT,
            duration_minutes INTEGER DEFAULT 60, total_points INTEGER DEFAULT 100,
            pass_score INTEGER DEFAULT 50, start_time TEXT, end_time TEXT,
            status TEXT DEFAULT 'draft', question_ids TEXT,
            randomize_questions INTEGER DEFAULT 0, randomize_options INTEGER DEFAULT 0,
            created_at TEXT DEFAULT 'now')''',
        f'''CREATE TABLE IF NOT EXISTS exam_access (
            id {AUTO}, exam_id INTEGER NOT NULL, student_id INTEGER NOT NULL,
            access_code TEXT UNIQUE NOT NULL, used INTEGER DEFAULT 0, used_at TEXT)''',
        f'''CREATE TABLE IF NOT EXISTS exam_sessions (
            id {AUTO}, exam_id INTEGER NOT NULL, student_id INTEGER NOT NULL,
            start_time TEXT DEFAULT 'now', end_time TEXT, status TEXT DEFAULT 'in_progress',
            time_remaining INTEGER)''',
        f'''CREATE TABLE IF NOT EXISTS student_answers (
            id {AUTO}, session_id INTEGER NOT NULL, question_id INTEGER NOT NULL,
            answer TEXT, is_correct INTEGER, points_earned REAL DEFAULT 0,
            teacher_grade REAL, teacher_feedback TEXT, answered_at TEXT DEFAULT 'now')''',
        f'''CREATE TABLE IF NOT EXISTS exam_results (
            id {AUTO}, session_id INTEGER NOT NULL, student_id INTEGER NOT NULL,
            exam_id INTEGER NOT NULL, total_score REAL DEFAULT 0, percentage REAL DEFAULT 0,
            grade_letter TEXT, passed INTEGER DEFAULT 0, submitted_at TEXT DEFAULT 'now')''',
    ]

    for sql in tables:
        if not USE_PG:
            sql = sql.replace('SERIAL PRIMARY KEY','INTEGER PRIMARY KEY AUTOINCREMENT')
        c.execute(sql)
    conn.commit()

    # Insert defaults only if not exists
    if USE_PG:
        c.execute("INSERT INTO users (username,password_hash,role,full_name_ar,full_name_en) VALUES (%s,%s,%s,%s,%s) ON CONFLICT(username) DO NOTHING",
                  ('admin', hash_pw('admin123'), 'admin', 'مدير النظام', 'System Admin'))
        for k,v in SCHOOL_CONFIG.items():
            c.execute("INSERT INTO school_settings VALUES (%s,%s) ON CONFLICT(key) DO NOTHING", (k,str(v)))
    else:
        c.execute("INSERT OR IGNORE INTO users (username,password_hash,role,full_name_ar,full_name_en) VALUES (?,?,?,?,?)",
                  ('admin', hash_pw('admin123'), 'admin', 'مدير النظام', 'System Admin'))
        for k,v in SCHOOL_CONFIG.items():
            c.execute("INSERT OR IGNORE INTO school_settings VALUES (?,?)", (k,str(v)))
    conn.commit(); conn.close()

# ── Auth ──────────────────────────────────────────────────────────────────────
def make_token(uid, role):
    return jwt.encode({'user_id':uid,'role':role,'exp':datetime.utcnow()+timedelta(hours=12)}, JWT_SECRET, algorithm='HS256')

def auth(roles=None):
    def dec(f):
        @wraps(f)
        def wrap(*a,**kw):
            token = request.headers.get('Authorization','').replace('Bearer ','')
            if not token: return jsonify({'error':'No token'}),401
            try:
                d = jwt.decode(token, JWT_SECRET, algorithms=['HS256'])
                if roles and d['role'] not in roles: return jsonify({'error':'Forbidden'}),403
                request.user = d; return f(*a,**kw)
            except jwt.ExpiredSignatureError: return jsonify({'error':'Token expired'}),401
            except: return jsonify({'error':'Invalid token'}),401
        return wrap
    return dec

# ── Excel Helpers ─────────────────────────────────────────────────────────────
def xl_response(wb, name):
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     download_name=name, as_attachment=True)

def xl_style(ws, headers, data, col_widths, title='', subtitle=''):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    H  = PatternFill("solid", fgColor="1a3a5c"); S = PatternFill("solid", fgColor="1e40af")
    E1 = PatternFill("solid", fgColor="f8fafc"); E2 = PatternFill("solid", fgColor="ffffff")
    thin = Side(style='thin', color='CBD5E1'); bdr = Border(left=thin,right=thin,top=thin,bottom=thin)
    row = 1
    if title:
        ws.merge_cells(f'A1:{get_column_letter(len(headers))}1')
        c=ws.cell(1,1,title); c.font=Font(name='Arial',bold=True,size=12,color='FFFFFF')
        c.fill=H; c.alignment=Alignment(horizontal='center',vertical='center'); ws.row_dimensions[1].height=28; row=2
    if subtitle:
        ws.merge_cells(f'A2:{get_column_letter(len(headers))}2')
        c=ws.cell(2,1,subtitle); c.font=Font(name='Arial',size=9,color='FFFFFF')
        c.fill=S; c.alignment=Alignment(horizontal='center',vertical='center'); ws.row_dimensions[2].height=18; row=3
    for ci,h in enumerate(headers,1):
        c=ws.cell(row,ci,h); c.font=Font(name='Arial',bold=True,size=10,color='FFFFFF')
        c.fill=S; c.alignment=Alignment(horizontal='center',vertical='center'); c.border=bdr
    ws.row_dimensions[row].height=22; row+=1
    for ri,rd in enumerate(data):
        fill = E1 if ri%2==0 else E2
        for ci,v in enumerate(rd,1):
            c=ws.cell(row,ci,v); c.font=Font(name='Arial',size=10)
            c.fill=fill; c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True); c.border=bdr
        ws.row_dimensions[row].height=20; row+=1
    for ci,w in enumerate(col_widths,1):
        ws.column_dimensions[get_column_letter(ci)].width=w
    ws.freeze_panes=f'A{(3 if subtitle else 2 if title else 1)+1}'

def read_xl_or_csv(file, key_col):
    fname = file.filename.lower(); rows=[]
    if fname.endswith('.xlsx') or fname.endswith('.xls'):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file.read()), data_only=True)
        ws = wb.active; hdrs=[]
        for row in ws.iter_rows(values_only=True):
            if all(v is None for v in row): continue
            if not hdrs:
                rv=[str(v).strip() if v else '' for v in row]
                if key_col in rv: hdrs=rv; continue
                else: continue
            rd={}
            for j,v in enumerate(row):
                if j<len(hdrs) and hdrs[j]: rd[hdrs[j]]=str(v).strip() if v is not None else ''
            if any(rd.get(h) for h in hdrs[:3]): rows.append(rd)
    else:
        content=file.read().decode('utf-8-sig')
        rows=[dict(r) for r in csv.DictReader(io.StringIO(content))]
    return rows

# ═══════════════════════════════════════════════════════════════
#  ROUTES
# ═══════════════════════════════════════════════════════════════

# ── Auth ──────────────────────────────────────────────────────
@app.route('/api/auth/login', methods=['POST'])
def login():
    d=request.json; conn=get_db()
    u=fetchone(conn,'SELECT * FROM users WHERE username=?',(d.get('username'),)); conn.close()
    if not u or u['password_hash']!=hash_pw(d.get('password','')): return jsonify({'error':'بيانات غير صحيحة'}),401
    return jsonify({'token':make_token(u['id'],u['role']),'role':u['role'],'name':u['full_name_ar'],'user_id':u['id']})

@app.route('/api/auth/me')
@auth()
def me():
    conn=get_db(); u=fetchone(conn,'SELECT id,username,role,full_name_ar,full_name_en,email,class_name,grade FROM users WHERE id=?',(request.user['user_id'],)); conn.close()
    return jsonify(u)

@app.route('/api/auth/change-password', methods=['POST'])
@auth()
def change_password():
    d=request.json; old=d.get('old_password',''); new=d.get('new_password','')
    if not new or len(new)<4: return jsonify({'error':'كلمة المرور قصيرة جداً'}),400
    conn=get_db(); u=fetchone(conn,'SELECT * FROM users WHERE id=?',(request.user['user_id'],))
    if not u or u['password_hash']!=hash_pw(old): conn.close(); return jsonify({'error':'كلمة المرور الحالية خاطئة'}),400
    execute(conn,'UPDATE users SET password_hash=? WHERE id=?',(hash_pw(new),request.user['user_id'])); conn.commit(); conn.close()
    return jsonify({'success':True})

# ── Settings ──────────────────────────────────────────────────
@app.route('/api/settings')
def get_settings():
    conn=get_db(); rows=fetchall(conn,'SELECT key,value FROM school_settings'); conn.close()
    return jsonify({r['key']:r['value'] for r in rows})

@app.route('/api/settings', methods=['PUT'])
@auth(['admin'])
def save_settings():
    conn=get_db()
    for k,v in request.json.items():
        if USE_PG: execute(conn,'INSERT INTO school_settings VALUES (?,?) ON CONFLICT(key) DO UPDATE SET value=EXCLUDED.value',(k,str(v)))
        else: execute(conn,'INSERT OR REPLACE INTO school_settings VALUES (?,?)',(k,str(v)))
    conn.commit(); conn.close(); return jsonify({'success':True})

# ── Users ─────────────────────────────────────────────────────
@app.route('/api/users')
@auth(['admin','teacher'])
def get_users():
    role=request.args.get('role',''); conn=get_db()
    if role: users=fetchall(conn,'SELECT id,username,role,full_name_ar,full_name_en,email,student_code,class_name,grade FROM users WHERE role=? ORDER BY full_name_ar',(role,))
    else: users=fetchall(conn,'SELECT id,username,role,full_name_ar,full_name_en,email,student_code,class_name,grade FROM users ORDER BY role,full_name_ar')
    conn.close(); return jsonify(users)

@app.route('/api/users', methods=['POST'])
@auth(['admin'])
def create_user():
    d=request.json; conn=get_db()
    code=gen_code(6) if d.get('role')=='student' else None
    try:
        insert(conn,'INSERT INTO users (username,password_hash,role,full_name_ar,full_name_en,email,student_code,class_name,grade) VALUES (?,?,?,?,?,?,?,?,?)',
               (d['username'],hash_pw(d.get('password','123456')),d['role'],d.get('full_name_ar'),d.get('full_name_en'),d.get('email'),code,d.get('class_name'),d.get('grade')))
        conn.commit(); conn.close(); return jsonify({'success':True,'student_code':code})
    except Exception as e: conn.close(); return jsonify({'error':str(e)}),400

@app.route('/api/users/<int:uid>', methods=['PUT'])
@auth(['admin'])
def update_user(uid):
    d=request.json; conn=get_db(); fields=[]; vals=[]
    for f in ['full_name_ar','full_name_en','email','class_name','grade','role']:
        if f in d: fields.append(f'{f}=?'); vals.append(d[f])
    if d.get('password'): fields.append('password_hash=?'); vals.append(hash_pw(d['password']))
    vals.append(uid); execute(conn,f'UPDATE users SET {",".join(fields)} WHERE id=?',vals)
    conn.commit(); conn.close(); return jsonify({'success':True})

@app.route('/api/users/<int:uid>', methods=['DELETE'])
@auth(['admin'])
def delete_user(uid):
    conn=get_db(); execute(conn,'DELETE FROM users WHERE id=?',(uid,)); conn.commit(); conn.close()
    return jsonify({'success':True})

@app.route('/api/users/import', methods=['POST'])
@auth(['admin'])
def import_users():
    file=request.files.get('file')
    if not file: return jsonify({'error':'No file'}),400
    role=request.args.get('role','student')
    try: rows=read_xl_or_csv(file,'username')
    except Exception as e: return jsonify({'error':str(e)}),400
    conn=get_db(); created=0; skipped=0; errors=[]
    for row in rows:
        try:
            uname=str(row.get('username') or '').strip()
            name_ar=str(row.get('full_name_ar') or '').strip()
            if not uname and not name_ar: continue
            if not uname: uname=f"u{gen_code(6).lower()}"
            if fetchone(conn,'SELECT id FROM users WHERE username=?',(uname,)): skipped+=1; continue
            code=gen_code(6) if role=='student' else None
            insert(conn,'INSERT INTO users (username,password_hash,role,full_name_ar,full_name_en,email,student_code,class_name,grade) VALUES (?,?,?,?,?,?,?,?,?)',
                   (uname,hash_pw(str(row.get('password') or '123456') or '123456'),role,name_ar,
                    str(row.get('full_name_en') or '').strip(),str(row.get('email') or '').strip(),
                    code,str(row.get('class_name') or '').strip(),str(row.get('grade') or '').strip()))
            created+=1
        except Exception as e: errors.append(f"{row.get('username','?')}: {str(e)}")
    conn.commit(); conn.close(); return jsonify({'created':created,'skipped':skipped,'errors':errors})

# ── Subjects ──────────────────────────────────────────────────
@app.route('/api/subjects')
@auth()
def get_subjects():
    conn=get_db()
    data=fetchall(conn,'SELECT s.*,u.full_name_ar as teacher_name FROM subjects s LEFT JOIN users u ON s.teacher_id=u.id ORDER BY s.name_ar')
    conn.close(); return jsonify(data)

@app.route('/api/subjects', methods=['POST'])
@auth(['admin'])
def create_subject():
    d=request.json; conn=get_db()
    try:
        insert(conn,'INSERT INTO subjects (name_ar,name_en,code,teacher_id,grade,color,icon) VALUES (?,?,?,?,?,?,?)',
               (d['name_ar'],d.get('name_en'),d['code'],d.get('teacher_id') or None,d.get('grade'),d.get('color','#3b82f6'),d.get('icon','📚')))
        conn.commit(); conn.close(); return jsonify({'success':True})
    except Exception as e: conn.close(); return jsonify({'error':str(e)}),400

@app.route('/api/subjects/<int:sid>', methods=['PUT'])
@auth(['admin'])
def update_subject(sid):
    d=request.json; conn=get_db()
    execute(conn,'UPDATE subjects SET name_ar=?,name_en=?,teacher_id=?,grade=?,color=?,icon=? WHERE id=?',
            (d['name_ar'],d.get('name_en'),d.get('teacher_id') or None,d.get('grade'),d.get('color'),d.get('icon'),sid))
    conn.commit(); conn.close(); return jsonify({'success':True})

@app.route('/api/subjects/<int:sid>', methods=['DELETE'])
@auth(['admin'])
def delete_subject(sid):
    conn=get_db(); execute(conn,'DELETE FROM subjects WHERE id=?',(sid,)); conn.commit(); conn.close()
    return jsonify({'success':True})

@app.route('/api/subjects/import', methods=['POST'])
@auth(['admin'])
def import_subjects():
    file=request.files.get('file')
    if not file: return jsonify({'error':'No file'}),400
    try: rows=read_xl_or_csv(file,'code')
    except Exception as e: return jsonify({'error':str(e)}),400
    conn=get_db(); created=0; skipped=0; errors=[]
    for row in rows:
        try:
            code=str(row.get('code') or '').strip(); name_ar=str(row.get('name_ar') or '').strip()
            if not code or not name_ar: continue
            if fetchone(conn,'SELECT id FROM subjects WHERE code=?',(code,)): skipped+=1; continue
            tid=None
            if row.get('teacher_username'):
                t=fetchone(conn,'SELECT id FROM users WHERE username=? AND role=?',(row['teacher_username'],'teacher'))
                if t: tid=t['id']
            insert(conn,'INSERT INTO subjects (name_ar,name_en,code,teacher_id,grade,color,icon) VALUES (?,?,?,?,?,?,?)',
                   (name_ar,str(row.get('name_en') or '').strip(),code,tid,
                    str(row.get('grade') or '').strip(),str(row.get('color') or '#3b82f6').strip(),str(row.get('icon') or '📚').strip()))
            created+=1
        except Exception as e: errors.append(f"{row.get('code','?')}: {str(e)}")
    conn.commit(); conn.close(); return jsonify({'created':created,'skipped':skipped,'errors':errors})

# ── Questions ─────────────────────────────────────────────────
@app.route('/api/questions')
@auth(['admin','teacher'])
def get_questions():
    sid=request.args.get('subject_id'); conn=get_db()
    if sid: data=fetchall(conn,'SELECT q.*,s.name_ar as subject_name FROM question_bank q JOIN subjects s ON q.subject_id=s.id WHERE q.subject_id=? ORDER BY q.id DESC',(sid,))
    else: data=fetchall(conn,'SELECT q.*,s.name_ar as subject_name FROM question_bank q JOIN subjects s ON q.subject_id=s.id ORDER BY q.id DESC')
    conn.close(); return jsonify(data)

@app.route('/api/questions', methods=['POST'])
@auth(['admin','teacher'])
def create_question():
    d=request.json; conn=get_db()
    insert(conn,'INSERT INTO question_bank (subject_id,question_ar,question_en,type,options_ar,options_en,correct_answer,points,difficulty,skill_tag,created_by) VALUES (?,?,?,?,?,?,?,?,?,?,?)',
           (d['subject_id'],d['question_ar'],d.get('question_en'),d['type'],
            json.dumps(d.get('options_ar',[]),ensure_ascii=False),json.dumps(d.get('options_en',[]),ensure_ascii=False),
            str(d.get('correct_answer','')),d.get('points',1),d.get('difficulty','medium'),d.get('skill_tag'),request.user['user_id']))
    conn.commit(); conn.close(); return jsonify({'success':True})

@app.route('/api/questions/<int:qid>', methods=['PUT'])
@auth(['admin','teacher'])
def update_question(qid):
    d=request.json; conn=get_db()
    execute(conn,'UPDATE question_bank SET subject_id=?,question_ar=?,question_en=?,type=?,options_ar=?,options_en=?,correct_answer=?,points=?,difficulty=?,skill_tag=? WHERE id=?',
            (d['subject_id'],d['question_ar'],d.get('question_en'),d['type'],
             json.dumps(d.get('options_ar',[]),ensure_ascii=False),json.dumps(d.get('options_en',[]),ensure_ascii=False),
             str(d.get('correct_answer','')),d.get('points',1),d.get('difficulty','medium'),d.get('skill_tag'),qid))
    conn.commit(); conn.close(); return jsonify({'success':True})

@app.route('/api/questions/<int:qid>', methods=['DELETE'])
@auth(['admin','teacher'])
def delete_question(qid):
    conn=get_db(); execute(conn,'DELETE FROM question_bank WHERE id=?',(qid,)); conn.commit(); conn.close()
    return jsonify({'success':True})

@app.route('/api/questions/import', methods=['POST'])
@auth(['admin','teacher'])
def import_questions():
    file=request.files.get('file')
    if not file: return jsonify({'error':'No file'}),400
    try: rows=read_xl_or_csv(file,'question_ar')
    except Exception as e: return jsonify({'error':str(e)}),400
    conn=get_db(); created=0; errors=[]
    for row in rows:
        try:
            q_ar=str(row.get('question_ar') or '').strip(); sc=str(row.get('subject_code') or '').strip()
            if not q_ar or not sc: continue
            subj=fetchone(conn,'SELECT id FROM subjects WHERE code=?',(sc,))
            if not subj: errors.append(f'مادة غير موجودة: {sc}'); continue
            qtype=str(row.get('type') or 'mcq').strip().lower()
            opts_ar=[]; opts_en=[]
            if qtype=='mcq':
                for i in range(1,6):
                    o=str(row.get(f'option{i}_ar') or row.get(f'option{i}') or '').strip()
                    oe=str(row.get(f'option{i}_en') or '').strip()
                    if o: opts_ar.append(o); opts_en.append(oe)
            insert(conn,'INSERT INTO question_bank (subject_id,question_ar,question_en,type,options_ar,options_en,correct_answer,points,difficulty,skill_tag,created_by) VALUES (?,?,?,?,?,?,?,?,?,?,?)',
                   (subj['id'],q_ar,str(row.get('question_en') or '').strip(),qtype,
                    json.dumps(opts_ar,ensure_ascii=False),json.dumps(opts_en,ensure_ascii=False),
                    str(row.get('correct_answer') or '').strip(),float(row.get('points') or 1),
                    str(row.get('difficulty') or 'medium').strip(),str(row.get('skill_tag') or '').strip(),
                    request.user['user_id'])); created+=1
        except Exception as e: errors.append(f"{row.get('question_ar','?')[:20]}: {str(e)}")
    conn.commit(); conn.close(); return jsonify({'created':created,'errors':errors})

# ── Exams ─────────────────────────────────────────────────────
@app.route('/api/exams')
@auth(['admin','teacher'])
def get_exams():
    conn=get_db(); uid=request.user['user_id']; role=request.user['role']
    if role=='admin': data=fetchall(conn,'SELECT e.*,s.name_ar as subject_name,u.full_name_ar as teacher_name FROM exams e JOIN subjects s ON e.subject_id=s.id JOIN users u ON e.teacher_id=u.id ORDER BY e.created_at DESC')
    else: data=fetchall(conn,'SELECT e.*,s.name_ar as subject_name,u.full_name_ar as teacher_name FROM exams e JOIN subjects s ON e.subject_id=s.id JOIN users u ON e.teacher_id=u.id WHERE e.teacher_id=? ORDER BY e.created_at DESC',(uid,))
    conn.close(); return jsonify(data)

@app.route('/api/exams', methods=['POST'])
@auth(['admin','teacher'])
def create_exam():
    d=request.json; conn=get_db()
    eid=insert(conn,'INSERT INTO exams (title_ar,title_en,subject_id,teacher_id,instructions_ar,instructions_en,duration_minutes,total_points,pass_score,start_time,end_time,question_ids,randomize_questions,randomize_options,status) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)',
               (d['title_ar'],d.get('title_en'),d['subject_id'],request.user['user_id'],d.get('instructions_ar'),d.get('instructions_en'),d.get('duration_minutes',60),d.get('total_points',100),d.get('pass_score',50),d.get('start_time'),d.get('end_time'),json.dumps(d.get('question_ids',[])),1 if d.get('randomize_questions') else 0,1 if d.get('randomize_options') else 0,d.get('status','draft')))
    conn.commit(); conn.close(); return jsonify({'success':True,'exam_id':eid})

@app.route('/api/exams/<int:eid>', methods=['PUT'])
@auth(['admin','teacher'])
def update_exam(eid):
    d=request.json; conn=get_db()
    execute(conn,'UPDATE exams SET title_ar=?,title_en=?,subject_id=?,instructions_ar=?,instructions_en=?,duration_minutes=?,total_points=?,pass_score=?,start_time=?,end_time=?,question_ids=?,randomize_questions=?,randomize_options=?,status=? WHERE id=?',
            (d['title_ar'],d.get('title_en'),d['subject_id'],d.get('instructions_ar'),d.get('instructions_en'),d.get('duration_minutes',60),d.get('total_points',100),d.get('pass_score',50),d.get('start_time'),d.get('end_time'),json.dumps(d.get('question_ids',[])),1 if d.get('randomize_questions') else 0,1 if d.get('randomize_options') else 0,d.get('status','draft'),eid))
    conn.commit(); conn.close(); return jsonify({'success':True})

@app.route('/api/exams/<int:eid>', methods=['DELETE'])
@auth(['admin','teacher'])
def delete_exam(eid):
    conn=get_db(); execute(conn,'DELETE FROM exams WHERE id=?',(eid,)); conn.commit(); conn.close()
    return jsonify({'success':True})

@app.route('/api/exams/<int:eid>/generate-codes', methods=['POST'])
@auth(['admin','teacher'])
def gen_codes(eid):
    sids=request.json.get('student_ids',[]); conn=get_db(); result=[]
    for sid in sids:
        ex=fetchone(conn,'SELECT access_code FROM exam_access WHERE exam_id=? AND student_id=?',(eid,sid))
        code=ex['access_code'] if ex else gen_code(8)
        if not ex: insert(conn,'INSERT INTO exam_access (exam_id,student_id,access_code) VALUES (?,?,?)',(eid,sid,code))
        st=fetchone(conn,'SELECT full_name_ar,class_name FROM users WHERE id=?',(sid,))
        result.append({'student_id':sid,'code':code,'name':st['full_name_ar'] if st else '','class':st['class_name'] if st else ''})
    conn.commit(); conn.close(); return jsonify(result)

# ── Exam Session ──────────────────────────────────────────────
@app.route('/api/exam/enter', methods=['POST'])
def enter_exam():
    code=request.json.get('code','').strip().upper(); conn=get_db()
    acc=fetchone(conn,'SELECT ea.*,e.title_ar,e.duration_minutes,e.status FROM exam_access ea JOIN exams e ON ea.exam_id=e.id WHERE ea.access_code=?',(code,))
    if not acc: conn.close(); return jsonify({'error':'كود غير صحيح'}),404
    if acc['used']: conn.close(); return jsonify({'error':'تم استخدام هذا الكود مسبقاً'}),400
    if acc['status']!='active': conn.close(); return jsonify({'error':'الامتحان غير متاح الآن'}),400
    sess=fetchone(conn,"SELECT * FROM exam_sessions WHERE exam_id=? AND student_id=? AND status='in_progress'",(acc['exam_id'],acc['student_id']))
    if sess: conn.close(); return jsonify({'session_id':sess['id'],'token':make_token(acc['student_id'],'student'),'exam_id':acc['exam_id'],'resume':True})
    sid=insert(conn,'INSERT INTO exam_sessions (exam_id,student_id,start_time) VALUES (?,?,?)',(acc['exam_id'],acc['student_id'],datetime.utcnow().isoformat()))
    execute(conn,'UPDATE exam_access SET used=1,used_at=? WHERE access_code=?',(datetime.utcnow().isoformat(),code))
    conn.commit(); conn.close()
    return jsonify({'session_id':sid,'token':make_token(acc['student_id'],'student'),'exam_id':acc['exam_id']})

@app.route('/api/exam/session/<int:sid>/questions')
@auth(['student'])
def session_questions(sid):
    conn=get_db()
    sess=fetchone(conn,'SELECT * FROM exam_sessions WHERE id=? AND student_id=?',(sid,request.user['user_id']))
    if not sess: conn.close(); return jsonify({'error':'Not found'}),404
    exam=fetchone(conn,'SELECT * FROM exams WHERE id=?',(sess['exam_id'],))
    qids=json.loads(exam['question_ids'] or '[]')
    if exam['randomize_questions']: random.shuffle(qids)
    qs=[]
    for qid in qids:
        q=fetchone(conn,'SELECT * FROM question_bank WHERE id=?',(qid,))
        if q:
            q['options_ar']=json.loads(q['options_ar'] or '[]'); q['options_en']=json.loads(q['options_en'] or '[]')
            q.pop('correct_answer',None); qs.append(q)
    start=datetime.fromisoformat(sess['start_time']); elapsed=(datetime.utcnow()-start).total_seconds()
    time_rem=max(0,exam['duration_minutes']*60-elapsed)
    answers=fetchall(conn,'SELECT question_id,answer FROM student_answers WHERE session_id=?',(sid,))
    conn.close()
    return jsonify({'exam':{'id':exam['id'],'title_ar':exam['title_ar'],'title_en':exam['title_en'],'duration_minutes':exam['duration_minutes'],'instructions_ar':exam['instructions_ar'],'instructions_en':exam['instructions_en'],'total_points':exam['total_points']},'questions':qs,'time_remaining':int(time_rem),'answers':{str(a['question_id']):a['answer'] for a in answers}})

@app.route('/api/exam/session/<int:sid>/answer', methods=['POST'])
@auth(['student'])
def save_answer(sid):
    d=request.json; conn=get_db()
    ex=fetchone(conn,'SELECT id FROM student_answers WHERE session_id=? AND question_id=?',(sid,d['question_id']))
    if ex: execute(conn,'UPDATE student_answers SET answer=? WHERE session_id=? AND question_id=?',(d['answer'],sid,d['question_id']))
    else: insert(conn,'INSERT INTO student_answers (session_id,question_id,answer) VALUES (?,?,?)',(sid,d['question_id'],d['answer']))
    conn.commit(); conn.close(); return jsonify({'success':True})

@app.route('/api/exam/session/<int:sid>/submit', methods=['POST'])
@auth(['student'])
def submit_exam(sid):
    conn=get_db()
    sess=fetchone(conn,'SELECT * FROM exam_sessions WHERE id=? AND student_id=?',(sid,request.user['user_id']))
    if not sess: conn.close(); return jsonify({'error':'Not found'}),404
    exam=fetchone(conn,'SELECT * FROM exams WHERE id=?',(sess['exam_id'],))
    answers=fetchall(conn,'SELECT * FROM student_answers WHERE session_id=?',(sid,))
    total=0
    for ans in answers:
        q=fetchone(conn,'SELECT * FROM question_bank WHERE id=?',(ans['question_id'],))
        if q and q['type']!='essay':
            correct=str(q['correct_answer']).strip().lower(); given=str(ans['answer'] or '').strip().lower()
            ok=1 if correct==given else 0; pts=q['points'] if ok else 0; total+=pts
            execute(conn,'UPDATE student_answers SET is_correct=?,points_earned=? WHERE id=?',(ok,pts,ans['id']))
    pct=(total/exam['total_points']*100) if exam['total_points']>0 else 0
    passed=1 if pct>=exam['pass_score'] else 0; gl=grade_letter(pct)
    execute(conn,'UPDATE exam_sessions SET status=?,end_time=? WHERE id=?',('submitted',datetime.utcnow().isoformat(),sid))
    if USE_PG: execute(conn,'INSERT INTO exam_results (session_id,student_id,exam_id,total_score,percentage,grade_letter,passed,submitted_at) VALUES (?,?,?,?,?,?,?,?) ON CONFLICT DO NOTHING',(sid,request.user['user_id'],sess['exam_id'],total,round(pct,1),gl,passed,datetime.utcnow().isoformat()))
    else: execute(conn,'INSERT OR REPLACE INTO exam_results (session_id,student_id,exam_id,total_score,percentage,grade_letter,passed,submitted_at) VALUES (?,?,?,?,?,?,?,?)',(sid,request.user['user_id'],sess['exam_id'],total,round(pct,1),gl,passed,datetime.utcnow().isoformat()))
    conn.commit(); conn.close()
    return jsonify({'score':total,'percentage':round(pct,1),'grade':gl,'passed':bool(passed)})

# ── Results & Analytics ────────────────────────────────────────
@app.route('/api/results/exam/<int:eid>')
@auth(['admin','teacher'])
def exam_results(eid):
    conn=get_db()
    data=fetchall(conn,'SELECT er.*,u.full_name_ar,u.class_name,u.grade,es.start_time,es.end_time FROM exam_results er JOIN users u ON er.student_id=u.id JOIN exam_sessions es ON er.session_id=es.id WHERE er.exam_id=? ORDER BY er.percentage DESC',(eid,))
    conn.close(); return jsonify(data)

@app.route('/api/results/student/<int:sid>')
@auth(['admin','teacher','student'])
def student_results(sid):
    if request.user['role']=='student' and request.user['user_id']!=sid: return jsonify({'error':'Forbidden'}),403
    conn=get_db()
    data=fetchall(conn,'SELECT er.*,e.title_ar,e.total_points,e.pass_score,s.name_ar as subject_name,s.color as subject_color FROM exam_results er JOIN exams e ON er.exam_id=e.id JOIN subjects s ON e.subject_id=s.id WHERE er.student_id=? ORDER BY er.submitted_at DESC',(sid,))
    conn.close(); return jsonify(data)

@app.route('/api/analytics/overview')
@auth(['admin','teacher'])
def analytics_overview():
    conn=get_db()
    stats={
        'total_students': fetchone(conn,"SELECT COUNT(*) as n FROM users WHERE role='student'")['n'],
        'total_teachers': fetchone(conn,"SELECT COUNT(*) as n FROM users WHERE role='teacher'")['n'],
        'total_exams':    fetchone(conn,"SELECT COUNT(*) as n FROM exams")['n'],
        'total_questions':fetchone(conn,"SELECT COUNT(*) as n FROM question_bank")['n'],
        'active_exams':   fetchone(conn,"SELECT COUNT(*) as n FROM exams WHERE status='active'")['n'],
        'total_submissions': fetchone(conn,"SELECT COUNT(*) as n FROM exam_results")['n'],
        'avg_score': 0, 'pass_rate': 0
    }
    r=fetchone(conn,"SELECT AVG(percentage) as avg, SUM(passed) as p, COUNT(*) as t FROM exam_results")
    if r and r['t']:
        stats['avg_score']=round(r['avg'] or 0,1); stats['pass_rate']=round((r['p'] or 0)/r['t']*100,1)
    stats['grade_distribution']=fetchall(conn,"SELECT grade_letter,COUNT(*) as count FROM exam_results GROUP BY grade_letter ORDER BY grade_letter")
    stats['subject_stats']=fetchall(conn,"SELECT s.name_ar,s.color,COUNT(er.id) as attempts,AVG(er.percentage) as avg_score,SUM(er.passed) as passed FROM subjects s LEFT JOIN exams e ON s.id=e.subject_id LEFT JOIN exam_results er ON e.id=er.exam_id GROUP BY s.id ORDER BY avg_score DESC")
    stats['recent_activity']=fetchall(conn,"SELECT er.*,u.full_name_ar,e.title_ar,s.name_ar as subject_name FROM exam_results er JOIN users u ON er.student_id=u.id JOIN exams e ON er.exam_id=e.id JOIN subjects s ON e.subject_id=s.id ORDER BY er.submitted_at DESC LIMIT 10")
    # Class comparison
    stats['class_comparison']=fetchall(conn,"SELECT u.class_name,COUNT(DISTINCT u.id) as students,COUNT(er.id) as attempts,AVG(er.percentage) as avg_score,SUM(er.passed) as passed FROM users u LEFT JOIN exam_results er ON u.id=er.student_id WHERE u.role='student' GROUP BY u.class_name ORDER BY avg_score DESC")
    # Skill stats
    stats['skill_stats']=fetchall(conn,"SELECT q.skill_tag,COUNT(sa.id) as attempts,SUM(sa.is_correct) as correct FROM question_bank q LEFT JOIN student_answers sa ON q.id=sa.question_id WHERE q.skill_tag IS NOT NULL AND q.skill_tag!='' GROUP BY q.skill_tag ORDER BY attempts DESC LIMIT 15")
    conn.close(); return jsonify(stats)

# ── PDF Reports ────────────────────────────────────────────────
def get_pdf_styles():
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib import colors
    return {
        'h1': ParagraphStyle('h1',fontName='Helvetica-Bold',fontSize=16,alignment=1,spaceAfter=6,textColor=colors.HexColor('#1a3a5c')),
        'h2': ParagraphStyle('h2',fontName='Helvetica-Bold',fontSize=12,spaceAfter=6,textColor=colors.HexColor('#1a3a5c')),
        'body': ParagraphStyle('body',fontName='Helvetica',fontSize=10,spaceAfter=4),
        'sub': ParagraphStyle('sub',fontName='Helvetica',fontSize=10,alignment=1,textColor=colors.HexColor('#64748b'),spaceAfter=4),
        'footer': ParagraphStyle('footer',fontName='Helvetica',fontSize=8,alignment=1,textColor=colors.HexColor('#94a3b8')),
    }

def build_pdf_header(story, settings, title, styles):
    from reportlab.platypus import Paragraph, Spacer, HRFlowable
    story.append(Paragraph(settings.get('name_en','School'), styles['h1']))
    story.append(Paragraph(settings.get('address',''), styles['sub']))
    story.append(HRFlowable(width='100%',thickness=2,color='#1a3a5c'))
    story.append(Spacer(1,0.3*72/2.54))
    story.append(Paragraph(title, styles['h1']))
    story.append(Spacer(1,0.3*72/2.54))

def pdf_table(data, col_widths, header_bg='#1a3a5c'):
    from reportlab.platypus import Table, TableStyle
    from reportlab.lib import colors
    t=Table(data,colWidths=col_widths)
    t.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,0),colors.HexColor(header_bg)),
        ('TEXTCOLOR',(0,0),(-1,0),colors.white),
        ('FONTNAME',(0,0),(-1,-1),'Helvetica'),
        ('FONTSIZE',(0,0),(-1,-1),9),
        ('GRID',(0,0),(-1,-1),0.5,colors.HexColor('#cbd5e1')),
        ('PADDING',(0,0),(-1,-1),6),
        ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,colors.HexColor('#f1f5f9')]),
    ]))
    return t

@app.route('/api/reports/student/<int:sid>')
@auth(['admin','teacher','student'])
def student_report(sid):
    if request.user['role']=='student' and request.user['user_id']!=sid: return jsonify({'error':'Forbidden'}),403
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    conn=get_db()
    student=fetchone(conn,'SELECT * FROM users WHERE id=?',(sid,))
    results=fetchall(conn,'SELECT er.*,e.title_ar,e.total_points,s.name_ar as subject_name FROM exam_results er JOIN exams e ON er.exam_id=e.id JOIN subjects s ON e.subject_id=s.id WHERE er.student_id=? ORDER BY er.submitted_at DESC',(sid,))
    skill_data=fetchall(conn,"SELECT q.skill_tag,COUNT(sa.id) as total,SUM(sa.is_correct) as correct FROM student_answers sa JOIN question_bank q ON sa.question_id=q.id JOIN exam_sessions es ON sa.session_id=es.id WHERE es.student_id=? AND q.skill_tag IS NOT NULL AND q.skill_tag!='' GROUP BY q.skill_tag",(sid,))
    settings={r['key']:r['value'] for r in fetchall(conn,'SELECT key,value FROM school_settings')}
    conn.close()
    buf=io.BytesIO()
    doc=SimpleDocTemplate(buf,pagesize=A4,rightMargin=2*cm,leftMargin=2*cm,topMargin=2*cm,bottomMargin=2*cm)
    story=[]; styles=get_pdf_styles()
    build_pdf_header(story,settings,f'تقرير الطالب / Student Report',styles)
    # Student info
    info=[['الاسم / Name',student.get('full_name_ar','')],['الفصل / Class',student.get('class_name','')],['المرحلة / Grade',student.get('grade','')],['تاريخ التقرير',datetime.now().strftime('%Y-%m-%d')]]
    story.append(pdf_table(info,[5*cm,10*cm]))
    story.append(Spacer(1,0.5*cm))
    if results:
        avg=sum(r['percentage'] for r in results)/len(results)
        passed=sum(1 for r in results if r['passed'])
        story.append(Paragraph('ملخص الأداء / Performance Summary', styles['h2']))
        story.append(pdf_table([['الامتحانات','المتوسط','ناجح','راسب'],[str(len(results)),f'{round(avg,1)}%',str(passed),str(len(results)-passed)]],[4.5*cm]*4))
        story.append(Spacer(1,0.5*cm))
        story.append(Paragraph('تفاصيل النتائج / Results Detail', styles['h2']))
        rows=[['الامتحان','المادة','الدرجة','التقدير','الحالة']]
        for r in results: rows.append([r['title_ar'][:30],r['subject_name'],f"{round(r['percentage'],1)}%",r['grade_letter'],'ناجح ✓' if r['passed'] else 'راسب ✗'])
        story.append(pdf_table(rows,[5*cm,3.5*cm,2.5*cm,2*cm,3*cm]))
    if skill_data:
        story.append(Spacer(1,0.5*cm))
        story.append(Paragraph('تقرير المهارات / Skills Report', styles['h2']))
        rows=[['المهارة','الإجابات','الصحيحة','نسبة الإتقان']]
        for sk in skill_data:
            pct=round((sk['correct'] or 0)/(sk['total'] or 1)*100,1)
            rows.append([sk['skill_tag'],str(sk['total']),str(sk['correct'] or 0),f'{pct}%'])
        story.append(pdf_table(rows,[5*cm,3*cm,3*cm,3*cm]))
    # Goal Setting Sheet
    story.append(Spacer(1,0.5*cm))
    story.append(Paragraph('هدفي / My Goal Setting', styles['h2']))
    goal_rows=[['المجال','الوضع الحالي','هدفي','خطة التحسين'],['المتوسط العام',f'{round(avg,1)}%' if results else '-','_____%','_________________'],['المادة الأضعف','_________','_____%','_________________'],['مهارة للتطوير','_________','_____%','_________________']]
    story.append(pdf_table(goal_rows,[4*cm,3*cm,3*cm,5.5*cm]))
    story.append(Spacer(1,1*cm))
    story.append(HRFlowable(width='100%',thickness=1,color='#cbd5e1'))
    story.append(Paragraph(settings.get('copyright',''), styles['footer']))
    doc.build(story); buf.seek(0)
    return send_file(buf,mimetype='application/pdf',download_name=f'report_{student.get("full_name_ar",sid)}.pdf')

@app.route('/api/reports/class/<class_name>')
@auth(['admin','teacher'])
def class_report(class_name):
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, HRFlowable
    from reportlab.lib.units import cm
    conn=get_db()
    students=fetchall(conn,"SELECT * FROM users WHERE class_name=? AND role='student' ORDER BY full_name_ar",(class_name,))
    settings={r['key']:r['value'] for r in fetchall(conn,'SELECT key,value FROM school_settings')}
    conn.close()
    buf=io.BytesIO()
    doc=SimpleDocTemplate(buf,pagesize=A4,rightMargin=2*cm,leftMargin=2*cm,topMargin=2*cm,bottomMargin=2*cm)
    story=[]; styles=get_pdf_styles(); conn=get_db()
    build_pdf_header(story,settings,f'تقرير الفصل: {class_name}',styles)
    rows=[['#','الطالب','الامتحانات','المتوسط','التقدير','ناجح/راسب']]
    for i,st in enumerate(students,1):
        res=fetchall(conn,'SELECT percentage,passed FROM exam_results WHERE student_id=?',(st['id'],))
        avg=sum(r['percentage'] for r in res)/len(res) if res else 0
        rows.append([str(i),st.get('full_name_ar',''),str(len(res)),f'{round(avg,1)}%',grade_letter(avg),f"{sum(1 for r in res if r['passed'])}/{len(res)}"])
    conn.close()
    story.append(pdf_table(rows,[0.8*cm,5.5*cm,2.5*cm,2.5*cm,2*cm,2.5*cm]))
    story.append(Spacer(1,1*cm)); story.append(HRFlowable(width='100%',thickness=1,color='#cbd5e1'))
    story.append(Paragraph(settings.get('copyright',''), styles['footer']))
    doc.build(story); buf.seek(0)
    return send_file(buf,mimetype='application/pdf',download_name=f'class_{class_name}.pdf')

@app.route('/api/reports/skills')
@auth(['admin','teacher'])
def skills_report():
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, HRFlowable
    from reportlab.lib.units import cm
    conn=get_db()
    settings={r['key']:r['value'] for r in fetchall(conn,'SELECT key,value FROM school_settings')}
    skill_data=fetchall(conn,"SELECT q.skill_tag,s.name_ar as subject,COUNT(sa.id) as attempts,SUM(sa.is_correct) as correct,COUNT(DISTINCT es.student_id) as students FROM question_bank q LEFT JOIN student_answers sa ON q.id=sa.question_id LEFT JOIN exam_sessions es ON sa.session_id=es.id JOIN subjects s ON q.subject_id=s.id WHERE q.skill_tag IS NOT NULL AND q.skill_tag!='' GROUP BY q.skill_tag,s.id ORDER BY s.name_ar,q.skill_tag")
    conn.close()
    buf=io.BytesIO()
    doc=SimpleDocTemplate(buf,pagesize=A4,rightMargin=2*cm,leftMargin=2*cm,topMargin=2*cm,bottomMargin=2*cm)
    story=[]; styles=get_pdf_styles()
    build_pdf_header(story,settings,'تقرير المهارات / Skills Report',styles)
    rows=[['المهارة','المادة','الطلاب','المحاولات','الصحيحة','نسبة الإتقان']]
    for sk in skill_data:
        pct=round((sk['correct'] or 0)/(sk['attempts'] or 1)*100,1)
        rows.append([sk['skill_tag'],sk['subject'],str(sk['students'] or 0),str(sk['attempts'] or 0),str(sk['correct'] or 0),f'{pct}%'])
    story.append(pdf_table(rows,[4*cm,3.5*cm,2*cm,2.5*cm,2.5*cm,2*cm]))
    story.append(Spacer(1,1*cm)); story.append(HRFlowable(width='100%',thickness=1,color='#cbd5e1'))
    story.append(Paragraph(settings.get('copyright',''), styles['footer']))
    doc.build(story); buf.seek(0)
    return send_file(buf,mimetype='application/pdf',download_name='skills_report.pdf')

# ── Excel Export ───────────────────────────────────────────────
@app.route('/api/export/<data_type>')
@auth(['admin','teacher'])
def export_excel(data_type):
    from openpyxl import Workbook
    conn=get_db()
    settings={r['key']:r['value'] for r in fetchall(conn,'SELECT key,value FROM school_settings')}
    school=settings.get('name_ar','مدرسة الرازي'); date_str=datetime.now().strftime('%Y-%m-%d %H:%M')
    wb=Workbook()

    if data_type=='students':
        data=fetchall(conn,"SELECT * FROM users WHERE role='student' ORDER BY class_name,full_name_ar")
        ws=wb.active; ws.title='الطلاب'; ws.sheet_view.rightToLeft=True
        xl_style(ws,['#','الاسم','اسم المستخدم','كود الطالب','الفصل','المرحلة','البريد'],
                 [[i+1,s.get('full_name_ar',''),s.get('username',''),s.get('student_code',''),s.get('class_name',''),s.get('grade',''),s.get('email','')] for i,s in enumerate(data)],
                 [5,28,18,14,10,12,28],f'{school} — الطلاب',f'تاريخ: {date_str} | العدد: {len(data)}')
        conn.close(); return xl_response(wb,f'students_{datetime.now().strftime("%Y%m%d")}.xlsx')

    elif data_type=='teachers':
        data=fetchall(conn,"SELECT * FROM users WHERE role='teacher' ORDER BY full_name_ar")
        ws=wb.active; ws.title='المعلمون'; ws.sheet_view.rightToLeft=True
        xl_style(ws,['#','الاسم','اسم المستخدم','البريد'],
                 [[i+1,t.get('full_name_ar',''),t.get('username',''),t.get('email','')] for i,t in enumerate(data)],
                 [5,28,18,28],f'{school} — المعلمون',f'تاريخ: {date_str}')
        conn.close(); return xl_response(wb,f'teachers_{datetime.now().strftime("%Y%m%d")}.xlsx')

    elif data_type=='subjects':
        data=fetchall(conn,'SELECT s.*,u.full_name_ar as teacher_name FROM subjects s LEFT JOIN users u ON s.teacher_id=u.id ORDER BY s.name_ar')
        ws=wb.active; ws.title='المواد'; ws.sheet_view.rightToLeft=True
        xl_style(ws,['#','المادة','Subject','الكود','المرحلة','المعلم'],
                 [[i+1,s.get('name_ar',''),s.get('name_en',''),s.get('code',''),s.get('grade',''),s.get('teacher_name','')] for i,s in enumerate(data)],
                 [5,22,22,12,12,22],f'{school} — المواد',f'تاريخ: {date_str}')
        conn.close(); return xl_response(wb,f'subjects_{datetime.now().strftime("%Y%m%d")}.xlsx')

    elif data_type=='questions':
        data=fetchall(conn,'SELECT q.*,s.name_ar as subject_name,s.code as subject_code FROM question_bank q JOIN subjects s ON q.subject_id=s.id ORDER BY s.name_ar')
        ws=wb.active; ws.title='الأسئلة'; ws.sheet_view.rightToLeft=True
        tm={'mcq':'اختيار متعدد','true_false':'صح/خطأ','essay':'مقالي'}
        dm={'easy':'سهل','medium':'متوسط','hard':'صعب'}
        rows=[]
        for i,q in enumerate(data):
            ans=q.get('correct_answer','')
            if q.get('type')=='mcq':
                try: opts=json.loads(q.get('options_ar') or '[]'); ans=opts[int(ans)] if int(ans)<len(opts) else ans
                except: pass
            elif q.get('type')=='true_false': ans='صح ✓' if ans=='true' else 'خطأ ✗'
            rows.append([i+1,q.get('subject_name',''),q.get('question_ar',''),tm.get(q.get('type',''),''),ans,q.get('points',1),dm.get(q.get('difficulty',''),''),q.get('skill_tag','')])
        xl_style(ws,['#','المادة','السؤال','النوع','الإجابة','الدرجة','الصعوبة','المهارة'],rows,
                 [5,18,40,14,20,8,10,14],f'{school} — بنك الأسئلة',f'تاريخ: {date_str}')
        conn.close(); return xl_response(wb,f'questions_{datetime.now().strftime("%Y%m%d")}.xlsx')

    elif data_type=='results':
        data=fetchall(conn,'SELECT er.*,u.full_name_ar,u.class_name,e.title_ar as exam_title,s.name_ar as subject_name FROM exam_results er JOIN users u ON er.student_id=u.id JOIN exams e ON er.exam_id=e.id JOIN subjects s ON e.subject_id=s.id ORDER BY u.class_name,u.full_name_ar')
        ws=wb.active; ws.title='النتائج'; ws.sheet_view.rightToLeft=True
        xl_style(ws,['#','الطالب','الفصل','الامتحان','المادة','الدرجة %','التقدير','الحالة'],
                 [[i+1,r.get('full_name_ar',''),r.get('class_name',''),r.get('exam_title',''),r.get('subject_name',''),f"{round(r.get('percentage',0),1)}%",r.get('grade_letter',''),'ناجح ✓' if r.get('passed') else 'راسب ✗'] for i,r in enumerate(data)],
                 [5,26,10,24,18,10,10,10],f'{school} — النتائج',f'تاريخ: {date_str} | العدد: {len(data)}')
        # Skills sheet
        ws2=wb.create_sheet('تحليل المهارات'); ws2.sheet_view.rightToLeft=True
        sk_data=fetchall(conn,"SELECT q.skill_tag,s.name_ar as subject,COUNT(sa.id) as attempts,SUM(sa.is_correct) as correct FROM question_bank q LEFT JOIN student_answers sa ON q.id=sa.question_id JOIN subjects s ON q.subject_id=s.id WHERE q.skill_tag IS NOT NULL AND q.skill_tag!='' GROUP BY q.skill_tag,s.id ORDER BY s.name_ar")
        xl_style(ws2,['المهارة','المادة','المحاولات','الصحيحة','نسبة الإتقان'],
                 [[sk.get('skill_tag',''),sk.get('subject',''),sk.get('attempts',0),sk.get('correct') or 0,f"{round((sk.get('correct') or 0)/(sk.get('attempts') or 1)*100,1)}%"] for sk in sk_data],
                 [20,18,12,12,14],f'{school} — تحليل المهارات',f'تاريخ: {date_str}')
        # Class comparison sheet
        ws3=wb.create_sheet('مقارنة الفصول'); ws3.sheet_view.rightToLeft=True
        cls_data=fetchall(conn,"SELECT u.class_name,COUNT(DISTINCT u.id) as students,COUNT(er.id) as attempts,AVG(er.percentage) as avg_score,SUM(er.passed) as passed FROM users u LEFT JOIN exam_results er ON u.id=er.student_id WHERE u.role='student' GROUP BY u.class_name ORDER BY avg_score DESC")
        xl_style(ws3,['الفصل','الطلاب','المحاولات','المتوسط','ناجح','معدل النجاح'],
                 [[c.get('class_name',''),c.get('students',0),c.get('attempts',0),f"{round(c.get('avg_score') or 0,1)}%",c.get('passed') or 0,f"{round((c.get('passed') or 0)/(c.get('attempts') or 1)*100,1)}%"] for c in cls_data],
                 [14,12,12,14,10,16],f'{school} — مقارنة الفصول',f'تاريخ: {date_str}')
        conn.close(); return xl_response(wb,f'full_report_{datetime.now().strftime("%Y%m%d")}.xlsx')

    conn.close(); return jsonify({'error':'Unknown type'}),400

# ── Backup & Restore ───────────────────────────────────────────
@app.route('/api/backup/download')
@auth(['admin'])
def backup_download():
    conn=get_db()
    backup={'version':'2.0','created_at':datetime.now().isoformat(),
            'school':{r['key']:r['value'] for r in fetchall(conn,'SELECT key,value FROM school_settings')},
            'users':fetchall(conn,'SELECT * FROM users'),
            'subjects':fetchall(conn,'SELECT * FROM subjects'),
            'questions':fetchall(conn,'SELECT * FROM question_bank'),
            'exams':fetchall(conn,'SELECT * FROM exams'),
            'exam_access':fetchall(conn,'SELECT * FROM exam_access'),
            'exam_sessions':fetchall(conn,'SELECT * FROM exam_sessions'),
            'student_answers':fetchall(conn,'SELECT * FROM student_answers'),
            'exam_results':fetchall(conn,'SELECT * FROM exam_results')}
    conn.close()
    buf=io.BytesIO(json.dumps(backup,ensure_ascii=False,indent=2).encode('utf-8')); buf.seek(0)
    return send_file(buf,mimetype='application/json',download_name=f'backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json',as_attachment=True)

@app.route('/api/backup/restore', methods=['POST'])
@auth(['admin'])
def backup_restore():
    file=request.files.get('file')
    if not file: return jsonify({'error':'No file'}),400
    try: data=json.loads(file.read().decode('utf-8'))
    except Exception as e: return jsonify({'error':str(e)}),400
    conn=get_db(); restored={}
    try:
        if 'school' in data:
            for k,v in data['school'].items():
                if USE_PG: execute(conn,'INSERT INTO school_settings VALUES (?,?) ON CONFLICT(key) DO UPDATE SET value=EXCLUDED.value',(k,v))
                else: execute(conn,'INSERT OR REPLACE INTO school_settings VALUES (?,?)',(k,v))
            restored['settings']=len(data['school'])
        for tbl,key,fields in [
            ('users','username','username,password_hash,role,full_name_ar,full_name_en,email,student_code,class_name,grade,created_at'),
            ('subjects','code','name_ar,name_en,code,teacher_id,grade,color,icon'),
        ]:
            count=0
            for row in data.get(tbl,[]):
                try:
                    vals=[row.get(f) for f in fields.split(',')]
                    if USE_PG: execute(conn,f"INSERT INTO {tbl} ({fields}) VALUES ({','.join(['?']*len(vals))}) ON CONFLICT({key}) DO NOTHING",vals)
                    else: execute(conn,f"INSERT OR IGNORE INTO {tbl} ({fields}) VALUES ({','.join(['?']*len(vals))})",vals)
                    count+=1
                except: pass
            restored[tbl]=count
        for row in data.get('questions',[]):
            try:
                execute(conn,'INSERT INTO question_bank (subject_id,question_ar,question_en,type,options_ar,options_en,correct_answer,points,difficulty,skill_tag,created_by,created_at) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)',
                        [row.get(f) for f in 'subject_id,question_ar,question_en,type,options_ar,options_en,correct_answer,points,difficulty,skill_tag,created_by,created_at'.split(',')])
            except: pass
        for row in data.get('exams',[]):
            try:
                execute(conn,'INSERT INTO exams (title_ar,title_en,subject_id,teacher_id,instructions_ar,instructions_en,duration_minutes,total_points,pass_score,start_time,end_time,status,question_ids,randomize_questions,randomize_options,created_at) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)',
                        [row.get(f) for f in 'title_ar,title_en,subject_id,teacher_id,instructions_ar,instructions_en,duration_minutes,total_points,pass_score,start_time,end_time,status,question_ids,randomize_questions,randomize_options,created_at'.split(',')])
            except: pass
        conn.commit(); conn.close(); return jsonify({'success':True,'restored':restored})
    except Exception as e: conn.close(); return jsonify({'error':str(e)}),500

# ── AI Question Generator ──────────────────────────────────────
@app.route('/api/ai/generate-questions', methods=['POST'])
@auth(['admin','teacher'])
def ai_generate():
    d=request.json
    conn=get_db()
    settings={r['key']:r['value'] for r in fetchall(conn,'SELECT key,value FROM school_settings')}
    conn.close()
    ai_key=settings.get('ai_key','') or os.environ.get('ANTHROPIC_KEY','')
    if not ai_key: return jsonify({'error':'يرجى إضافة مفتاح Anthropic API في إعدادات المدرسة'}),400
    import urllib.request, urllib.error
    prompt=f"""أنت مساعد تعليمي. أنشئ {d.get('count',5)} أسئلة {d.get('type','mcq')} باللغة العربية عن موضوع: "{d.get('topic','')}"
للمرحلة: {d.get('grade','')} | المادة: {d.get('subject','')} | الصعوبة: {d.get('difficulty','medium')}
أرجع JSON فقط بهذا الشكل بدون أي نص آخر:
{{"questions":[{{"question_ar":"...","question_en":"...","type":"{d.get('type','mcq')}","options_ar":["أ","ب","ج","د"],"options_en":["A","B","C","D"],"correct_answer":"0","points":1,"difficulty":"{d.get('difficulty','medium')}","skill_tag":"{d.get('skill','')}","explanation":"..."}}]}}
للصح/خطأ: options_ar=[] و correct_answer="true" أو "false"
للمقالي: options_ar=[] و correct_answer="" """
    try:
        req_data=json.dumps({'model':'claude-sonnet-4-20250514','max_tokens':2000,'messages':[{'role':'user','content':prompt}]}).encode()
        req=urllib.request.Request('https://api.anthropic.com/v1/messages',data=req_data,
            headers={'Content-Type':'application/json','x-api-key':ai_key,'anthropic-version':'2023-06-01'})
        with urllib.request.urlopen(req,timeout=30) as resp:
            result=json.loads(resp.read())
        text=result['content'][0]['text'].strip()
        if text.startswith('```'): text=text.split('```')[1]; text=text[text.find('{'):]
        questions=json.loads(text)['questions']
        return jsonify({'success':True,'questions':questions})
    except Exception as e: return jsonify({'error':f'خطأ في الذكاء الاصطناعي: {str(e)}'}),500

# ── Serve Frontend ─────────────────────────────────────────────
@app.route('/', defaults={'path':''})
@app.route('/<path:path>')
def serve(path):
    if path and os.path.exists(f'public/{path}'): return send_file(f'public/{path}')
    return send_file('public/index.html')

if __name__=='__main__':
    init_db()
    port=int(os.environ.get('PORT',5000))
    print(f"\n{'='*55}\n  مدرسة الرازي — نظام الامتحانات الذكي v2.0\n  http://localhost:{port}\n  admin / admin123\n{'='*55}\n")
    app.run(debug=False,host='0.0.0.0',port=port)
